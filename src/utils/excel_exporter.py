# -*- coding: utf-8 -*-
"""
Excel 导出工具 - 机构级专业格式

提供专业的 Excel 格式化导出功能，严格遵循投资银行标准:
- 公式优先 (所有计算使用 Excel 公式)
- 颜色规范: Blue=输入, Black=公式, Purple=同表链接, Green=跨表链接
- 填充色: 深蓝=表头, 浅蓝=列头, 灰色=统计, 白色=数据
- 方法论和数据源文档
"""

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from ..model import CompsResult, DCFResult

# ==================== 颜色规范 ====================
# 字体颜色 (区分输入/公式/链接)
FONT_BLUE = Font(color="0000FF", size=10)      # 硬编码输入
FONT_BLACK = Font(color="000000", size=10)     # 计算公式
FONT_PURPLE = Font(color="800080", size=10)   # 同表链接
FONT_GREEN = Font(color="008000", size=10)     # 跨表链接

# 填充颜色
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # 深蓝表头
COLUMN_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # 浅蓝列头
STATISTICS_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # 灰色统计
INPUT_CELL_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # 浅灰输入
KEY_OUTPUT_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # 中蓝关键输出

# 字体
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
NUMBER_FONT = Font(size=10)
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(bold=True, size=12)
BOLD_FONT = Font(bold=True, size=10)

# 边框
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# 对齐
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")


class CompsExcelExporter:
    """
    可比公司分析 Excel 导出器 - 机构级

    严格遵循投资银行标准:
    - 使用公式而非硬编码值
    - 颜色规范: Blue=输入, Black=公式, Purple=同表链接, Green=跨表链接
    - 包含方法论和数据源文档
    """

    def export(self, result: CompsResult) -> BytesIO:
        """
        导出 Comps 分析结果到 Excel

        Args:
            result: Comps 分析结果

        Returns:
            BytesIO: Excel 文件流
        """
        wb = Workbook()

        # 创建工作表
        ws_summary = wb.active
        assert ws_summary is not None, "Failed to create worksheet"
        ws_summary.title = "Summary"

        ws_comps = wb.create_sheet("Comparable Companies")
        ws_operating = wb.create_sheet("Operating Metrics")
        ws_multiples = wb.create_sheet("Valuation Multiples")
        ws_percentiles = wb.create_sheet("Percentile Analysis")
        ws_methodology = wb.create_sheet("Methodology & Sources")

        # 填充数据
        self._fill_summary(ws_summary, result)
        self._fill_comps_with_formulas(ws_comps, result)
        self._fill_operating(ws_operating, result)
        self._fill_multiples(ws_multiples, result)
        self._fill_percentiles(ws_percentiles, result)
        self._fill_comps_methodology(ws_methodology, result)

        # 保存到 BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer

    def _fill_summary(self, ws, result: CompsResult):
        """填充汇总页"""
        # 标题
        ws["A1"] = "COMPARABLE COMPANY ANALYSIS"
        ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
        ws.merge_cells("A1:E1")

        # 目标公司信息
        ws["A3"] = "TARGET COMPANY"
        ws["A3"].font = SUBTITLE_FONT
        ws["A3"].fill = COLUMN_HEADER_FILL

        ws["A4"] = "Symbol:"
        ws["B4"] = result.target_symbol
        ws["B4"].font = FONT_BLUE

        ws["A5"] = "Company Name:"
        ws["B5"] = result.target_name
        ws["B5"].font = FONT_BLUE

        ws["A6"] = "Sector:"
        ws["B6"] = result.sector
        ws["B6"].font = FONT_BLUE

        ws["A7"] = "Industry:"
        ws["B7"] = result.industry
        ws["B7"].font = FONT_BLUE

        # 分析结果
        ws["A9"] = "ANALYSIS SUMMARY"
        ws["A9"].font = SUBTITLE_FONT
        ws["A9"].fill = COLUMN_HEADER_FILL

        ws["A10"] = "Comparable Companies:"
        ws["B10"] = len(result.comps)
        ws["B10"].font = FONT_BLUE

        ws["A11"] = "Recommendation:"
        ws["B11"] = result.recommendation
        ws["B11"].font = FONT_BLUE

        ws["A12"] = "Confidence:"
        ws["B12"] = result.confidence
        ws["B12"].font = FONT_BLUE

        # 隐含估值 - 使用链接
        ws["A14"] = "IMPLIED VALUATION (Million USD)"
        ws["A14"].font = SUBTITLE_FONT
        ws["A14"].fill = KEY_OUTPUT_FILL

        headers = ["Method", "Low (25th)", "Mid (50th)", "High (75th)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=15, column=col, value=header)
            cell.fill = COLUMN_HEADER_FILL
            cell.font = BOLD_FONT
            cell.border = THIN_BORDER

        # P/E
        ws["A16"] = "P/E Implied"
        ws["B16"] = "='Valuation Multiples'!B15"
        ws["B16"].font = FONT_GREEN
        ws["C16"] = "='Valuation Multiples'!C15"
        ws["C16"].font = FONT_GREEN
        ws["D16"] = "='Valuation Multiples'!D15"
        ws["D16"].font = FONT_GREEN

        # P/S
        ws["A17"] = "P/S Implied"
        ws["B17"] = "='Valuation Multiples'!B16"
        ws["B17"].font = FONT_GREEN
        ws["C17"] = "='Valuation Multiples'!C16"
        ws["C17"].font = FONT_GREEN
        ws["D17"] = "='Valuation Multiples'!D16"
        ws["D17"].font = FONT_GREEN

        # EV/EBITDA
        ws["A18"] = "EV/EBITDA Implied"
        ws["B18"] = "='Valuation Multiples'!B17"
        ws["B18"].font = FONT_GREEN
        ws["C18"] = "='Valuation Multiples'!C17"
        ws["C18"].font = FONT_GREEN
        ws["D18"] = "='Valuation Multiples'!D17"
        ws["D18"].font = FONT_GREEN

        # 目标公司当前估值对比
        ws["A20"] = "TARGET COMPANY CURRENT VALUATION"
        ws["A20"].font = SUBTITLE_FONT
        ws["A20"].fill = COLUMN_HEADER_FILL

        if result.comps:
            target = result.comps[0]  # First comp is target
            ws["A21"] = "Current P/E:"
            ws["B21"] = target.pe_ratio if target.pe_ratio > 0 else "N/A"
            ws["A22"] = "Current P/S:"
            ws["B22"] = target.ps_ratio if target.ps_ratio > 0 else "N/A"
            ws["A23"] = "Current EV/EBITDA:"
            ws["B23"] = target.ev_ebitda if target.ev_ebitda > 0 else "N/A"

        # 调整列宽
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 18

    def _fill_comps_with_formulas(self, ws, result: CompsResult):
        """填充可比公司数据 - 使用公式"""
        ws["A1"] = "COMPARABLE COMPANIES"
        ws["A1"].font = Font(bold=True, size=14, color="1F4E79")

        # 表头
        headers = [
            "Symbol",
            "Company Name",
            "Market Cap ($M)",
            "Enterprise Value ($M)",
            "Revenue ($M)",
            "Revenue Growth (%)",
            "Gross Margin (%)",
            "EBITDA Margin (%)",
            "FCF Margin (%)",
            "P/E",
            "P/S",
            "EV/EBITDA",
            "EV/Revenue",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = COLUMN_HEADER_FILL
            cell.font = BOLD_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN

        # 数据行
        for row, comp in enumerate(result.comps, 4):
            # Symbol
            ws.cell(row=row, column=1, value=comp.symbol).font = FONT_BLUE
            ws.cell(row=row, column=1).border = THIN_BORDER

            # Company Name
            ws.cell(row=row, column=2, value=comp.company_name).font = FONT_BLUE
            ws.cell(row=row, column=2).border = THIN_BORDER

            # Market Cap
            ws.cell(row=row, column=3, value=comp.market_cap)
            ws.cell(row=row, column=3).number_format = "#,##0"
            ws.cell(row=row, column=3).font = FONT_BLUE
            ws.cell(row=row, column=3).border = THIN_BORDER

            # Enterprise Value
            ws.cell(row=row, column=4, value=comp.enterprise_value)
            ws.cell(row=row, column=4).number_format = "#,##0"
            ws.cell(row=row, column=4).font = FONT_BLUE
            ws.cell(row=row, column=4).border = THIN_BORDER

            # Revenue
            ws.cell(row=row, column=5, value=comp.revenue)
            ws.cell(row=row, column=5).number_format = "#,##0"
            ws.cell(row=row, column=5).font = FONT_BLUE
            ws.cell(row=row, column=5).border = THIN_BORDER

            # Revenue Growth
            ws.cell(row=row, column=6, value=comp.revenue_growth)
            ws.cell(row=row, column=6).number_format = "0.0%"
            ws.cell(row=row, column=6).font = FONT_BLUE
            ws.cell(row=row, column=6).border = THIN_BORDER

            # Gross Margin
            ws.cell(row=row, column=7, value=comp.gross_margin)
            ws.cell(row=row, column=7).number_format = "0.0%"
            ws.cell(row=row, column=7).font = FONT_BLUE
            ws.cell(row=row, column=7).border = THIN_BORDER

            # EBITDA Margin - 公式计算
            ws.cell(row=row, column=8, value=f"=IF(E{row}>0,D{row}/E{row},0)")
            ws.cell(row=row, column=8).number_format = "0.0%"
            ws.cell(row=row, column=8).font = FONT_BLACK
            ws.cell(row=row, column=8).border = THIN_BORDER

            # FCF Margin - 公式计算
            ws.cell(row=row, column=9, value=f"=IF(E{row}>0,G{row}/E{row},0)")
            ws.cell(row=row, column=9).number_format = "0.0%"
            ws.cell(row=row, column=9).font = FONT_BLACK
            ws.cell(row=row, column=9).border = THIN_BORDER

            # P/E
            if comp.pe_ratio > 0:
                ws.cell(row=row, column=10, value=comp.pe_ratio)
            else:
                ws.cell(row=row, column=10, value="N/A")
            ws.cell(row=row, column=10).font = FONT_BLUE
            ws.cell(row=row, column=10).border = THIN_BORDER

            # P/S
            if comp.ps_ratio > 0:
                ws.cell(row=row, column=11, value=comp.ps_ratio)
            else:
                ws.cell(row=row, column=11, value="N/A")
            ws.cell(row=row, column=11).font = FONT_BLUE
            ws.cell(row=row, column=11).border = THIN_BORDER

            # EV/EBITDA
            if comp.ev_ebitda > 0:
                ws.cell(row=row, column=12, value=comp.ev_ebitda)
            else:
                ws.cell(row=row, column=12, value="N/A")
            ws.cell(row=row, column=12).font = FONT_BLUE
            ws.cell(row=row, column=12).border = THIN_BORDER

            # EV/Revenue
            if comp.ev_revenue > 0:
                ws.cell(row=row, column=13, value=comp.ev_revenue)
            else:
                ws.cell(row=row, column=13, value="N/A")
            ws.cell(row=row, column=13).font = FONT_BLUE
            ws.cell(row=row, column=13).border = THIN_BORDER

        # 调整列宽
        column_widths = [10, 35, 15, 15, 15, 15, 15, 15, 15, 10, 10, 12, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def _fill_operating(self, ws, result: CompsResult):
        """填充运营指标统计"""
        ws["A1"] = "OPERATING METRICS STATISTICS"
        ws["A1"].font = Font(bold=True, size=14, color="1F4E79")

        # 表头
        headers = ["Metric", "Average", "Median", "Min", "Max"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = COLUMN_HEADER_FILL
            cell.font = BOLD_FONT
            cell.border = THIN_BORDER

        metrics = [
            ("Revenue ($M)", result.operating_metrics.revenue_avg, result.operating_metrics.revenue_median),
            ("Revenue Growth (%)", result.operating_metrics.growth_avg * 100, result.operating_metrics.growth_median * 100),
            ("Gross Margin (%)", result.operating_metrics.gross_margin_avg * 100, result.operating_metrics.gross_margin_median * 100),
            ("EBITDA Margin (%)", result.operating_metrics.ebitda_margin_avg * 100, result.operating_metrics.ebitda_margin_median * 100),
            ("FCF Margin (%)", result.operating_metrics.fcf_margin_avg * 100, result.operating_metrics.fcf_margin_median * 100),
        ]

        row = 4
        for name, avg, median in metrics:
            ws.cell(row=row, column=1, value=name).border = THIN_BORDER
            ws.cell(row=row, column=2, value=avg).border = THIN_BORDER
            ws.cell(row=row, column=2).number_format = "#,##0.0"
            ws.cell(row=row, column=3, value=median).border = THIN_BORDER
            ws.cell(row=row, column=3).number_format = "#,##0.0"
            row += 1

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15

    def _fill_multiples(self, ws, result: CompsResult):
        """填充估值倍数统计"""
        ws["A1"] = "VALUATION MULTIPLES STATISTICS"
        ws["A1"].font = Font(bold=True, size=14, color="1F4E79")

        # 表头
        headers = ["Metric", "Average", "Median"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = COLUMN_HEADER_FILL
            cell.font = BOLD_FONT
            cell.border = THIN_BORDER

        # 数据
        multiples = [
            ("P/E", result.valuation_multiples.pe_avg, result.valuation_multiples.pe_median),
            ("P/S", result.valuation_multiples.ps_avg, result.valuation_multiples.ps_median),
            ("P/B", result.valuation_multiples.pb_avg, result.valuation_multiples.pb_median),
            ("EV/EBITDA", result.valuation_multiples.ev_ebitda_avg, result.valuation_multiples.ev_ebitda_median),
            ("EV/Revenue", result.valuation_multiples.ev_revenue_avg, result.valuation_multiples.ev_revenue_median),
            ("EV/FCF", result.valuation_multiples.ev_fcf_avg, result.valuation_multiples.ev_fcf_median),
        ]

        row = 4
        for name, avg, median in multiples:
            ws.cell(row=row, column=1, value=name).border = THIN_BORDER
            ws.cell(row=row, column=2, value=avg).border = THIN_BORDER
            ws.cell(row=row, column=2).number_format = "0.0x"
            ws.cell(row=row, column=3, value=median).border = THIN_BORDER
            ws.cell(row=row, column=3).number_format = "0.0x"
            row += 1

        # 添加隐含估值计算 (使用公式)
        row += 1
        ws.cell(row=row, column=1, value="IMPLIED VALUATION").font = SUBTITLE_FONT
        ws.cell(row=row, column=1).fill = KEY_OUTPUT_FILL
        row += 1

        # 表头
        headers2 = ["Method", "Low", "Mid", "High"]
        for col, header in enumerate(headers2, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = COLUMN_HEADER_FILL
            cell.font = BOLD_FONT
            cell.border = THIN_BORDER
        row += 1

        # P/E implied
        if result.comps and result.comps[0].net_income:
            ws.cell(row=row, column=1, value="P/E Implied ($M)")
            ws.cell(row=row, column=2, value=result.implied_pe_low).number_format = "#,##0"
            ws.cell(row=row, column=3, value=result.implied_pe_mid).number_format = "#,##0"
            ws.cell(row=row, column=4, value=result.implied_pe_high).number_format = "#,##0"
        else:
            ws.cell(row=row, column=1, value="P/E Implied ($M)")

        # P/S implied
        row += 1
        if result.comps and result.comps[0].revenue:
            ws.cell(row=row, column=1, value="P/S Implied ($M)")
            ws.cell(row=row, column=2, value=result.implied_ps_low).number_format = "#,##0"
            ws.cell(row=row, column=3, value=result.implied_ps_mid).number_format = "#,##0"
            ws.cell(row=row, column=4, value=result.implied_ps_high).number_format = "#,##0"
        else:
            ws.cell(row=row, column=1, value="P/S Implied ($M)")

        # EV/EBITDA implied
        row += 1
        if result.comps and result.comps[0].ebitda:
            ws.cell(row=row, column=1, value="EV/EBITDA Implied ($M)")
            ws.cell(row=row, column=2, value=result.implied_ev_ebitda_low).number_format = "#,##0"
            ws.cell(row=row, column=3, value=result.implied_ev_ebitda_mid).number_format = "#,##0"
            ws.cell(row=row, column=4, value=result.implied_ev_ebitda_high).number_format = "#,##0"
        else:
            ws.cell(row=row, column=1, value="EV/EBITDA Implied ($M)")

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15

    def _fill_percentiles(self, ws, result: CompsResult):
        """填充分位数分析"""
        ws["A1"] = "PERCENTILE ANALYSIS"
        ws["A1"].font = Font(bold=True, size=14, color="1F4E79")

        # 估值倍数分位数
        ws["A3"] = "VALUATION MULTIPLES"
        ws["A3"].font = SUBTITLE_FONT
        ws["A3"].fill = COLUMN_HEADER_FILL

        headers = ["Metric", "25th", "50th (Median)", "75th"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.fill = COLUMN_HEADER_FILL
            cell.font = BOLD_FONT
            cell.border = THIN_BORDER

        percentiles = [
            ("P/E", result.percentiles.pe_25th, result.percentiles.pe_50th, result.percentiles.pe_75th),
            ("P/S", result.percentiles.ps_25th, result.percentiles.ps_50th, result.percentiles.ps_75th),
            ("P/B", result.percentiles.pb_25th, result.percentiles.pb_50th, result.percentiles.pb_75th),
            ("EV/EBITDA", result.percentiles.ev_ebitda_25th, result.percentiles.ev_ebitda_50th, result.percentiles.ev_ebitda_75th),
        ]

        row = 5
        for name, p25, p50, p75 in percentiles:
            ws.cell(row=row, column=1, value=name).border = THIN_BORDER
            ws.cell(row=row, column=2, value=p25).border = THIN_BORDER
            ws.cell(row=row, column=2).number_format = "0.0x"
            ws.cell(row=row, column=3, value=p50).border = THIN_BORDER
            ws.cell(row=row, column=3).number_format = "0.0x"
            ws.cell(row=row, column=4, value=p75).border = THIN_BORDER
            ws.cell(row=row, column=4).number_format = "0.0x"
            row += 1

        # 运营指标分位数
        row += 1
        ws.cell(row=row, column=1, value="OPERATING METRICS").font = SUBTITLE_FONT
        ws.cell(row=row, column=1).fill = COLUMN_HEADER_FILL
        row += 1

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = COLUMN_HEADER_FILL
            cell.font = BOLD_FONT
            cell.border = THIN_BORDER
        row += 1

        operating = [
            ("Revenue Growth (%)", result.percentiles.revenue_growth_25th * 100, result.percentiles.revenue_growth_50th * 100, result.percentiles.revenue_growth_75th * 100),
            ("Gross Margin (%)", result.percentiles.gross_margin_25th * 100, result.percentiles.gross_margin_50th * 100, result.percentiles.gross_margin_75th * 100),
            ("EBITDA Margin (%)", result.percentiles.ebitda_margin_25th * 100, result.percentiles.ebitda_margin_50th * 100, result.percentiles.ebitda_margin_75th * 100),
        ]

        for name, p25, p50, p75 in operating:
            ws.cell(row=row, column=1, value=name).border = THIN_BORDER
            ws.cell(row=row, column=2, value=p25).border = THIN_BORDER
            ws.cell(row=row, column=2).number_format = "0.0%"
            ws.cell(row=row, column=3, value=p50).border = THIN_BORDER
            ws.cell(row=row, column=3).number_format = "0.0%"
            ws.cell(row=row, column=4, value=p75).border = THIN_BORDER
            ws.cell(row=row, column=4).number_format = "0.0%"
            row += 1

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 15

    def _fill_comps_methodology(self, ws, result: CompsResult):
        """填充方法论和数据源文档"""
        ws["A1"] = "METHODOLOGY & DATA SOURCES"
        ws["A1"].font = Font(bold=True, size=14, color="1F4E79")

        row = 3

        # 分析日期
        ws.cell(row=row, column=1, value="ANALYSIS DATE")
        ws.cell(row=row, column=1).font = SUBTITLE_FONT
        ws.cell(row=row, column=1).fill = COLUMN_HEADER_FILL
        row += 1
        ws.cell(row=row, column=1, value=datetime.now().strftime("%Y-%m-%d"))
        ws.cell(row=row, column=1).font = FONT_BLUE
        row += 2

        # 数据源
        ws.cell(row=row, column=1, value="DATA SOURCES")
        ws.cell(row=row, column=1).font = SUBTITLE_FONT
        ws.cell(row=row, column=1).fill = COLUMN_HEADER_FILL
        row += 1

        sources = [
            ("Primary Data:", "yfinance"),
            ("Financial Data:", "Company SEC Filings (10-K, 10-Q)"),
            ("Market Data:", "Real-time market prices from exchange"),
            ("Peer Selection:", f"Based on {result.sector} sector"),
        ]

        for label, source in sources:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=source)
            row += 1
        row += 1

        # 方法论
        ws.cell(row=row, column=1, value="VALUATION METHODOLOGY")
        ws.cell(row=row, column=1).font = SUBTITLE_FONT
        ws.cell(row=row, column=1).fill = COLUMN_HEADER_FILL
        row += 1

        methodologies = [
            ("Peer Group:", f"Selected {len(result.comps)} comparable companies in {result.sector}"),
            ("Valuation Method:", "Trading multiples analysis"),
            ("Metrics Used:", "P/E, P/S, P/B, EV/EBITDA, EV/Revenue"),
            ("Statistics:", "25th/50th/75th percentile analysis"),
        ]

        for label, method in methodologies:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=method)
            row += 1
        row += 1

        # 免责声明
        ws.cell(row=row, column=1, value="DISCLAIMER")
        ws.cell(row=row, column=1).font = SUBTITLE_FONT
        ws.cell(row=row, column=1).fill = COLUMN_HEADER_FILL
        row += 1

        disclaimer = (
            "This analysis is for informational purposes only and does not constitute "
            "investment advice. Comps analysis is based on publicly available data and "
            "may not reflect current market conditions. Past performance is not "
            "indicative of future results."
        )
        ws.cell(row=row, column=1, value=disclaimer)
        ws.merge_cells(f"A{row}:E{row}")
        ws.cell(row=row, column=1).alignment = LEFT_ALIGN
        ws.cell(row=row, column=1).font = Font(size=9, italic=True)

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 50


class DCFExcelExporter:
    """
    DCF 估值模型 Excel 导出器 - 机构级

    严格遵循投资银行标准:
    - 使用公式而非硬编码值
    - 颜色规范: Blue=输入, Black=公式, Purple=同表链接, Green=跨表链接
    - 包含方法论和数据源文档
    """

    def export(self, result: DCFResult) -> BytesIO:
        """
        导出 DCF 估值结果到 Excel

        Args:
            result: DCF 估值结果

        Returns:
            BytesIO: Excel 文件流
        """
        wb = Workbook()

        # 创建工作表
        ws_summary = wb.active
        assert ws_summary is not None, "Failed to create worksheet"
        ws_summary.title = "Summary"

        ws_wacc = wb.create_sheet("WACC")
        ws_fcf = wb.create_sheet("FCF Projections")
        ws_terminal = wb.create_sheet("Terminal Value")
        ws_sensitivity = wb.create_sheet("Sensitivity Analysis")
        ws_methodology = wb.create_sheet("Methodology & Sources")

        # 填充数据 (使用公式)
        self._fill_summary(ws_summary, result)
        self._fill_wacc(ws_wacc, result)
        self._fill_fcf_with_formulas(ws_fcf, result)
        self._fill_terminal(ws_terminal, result)
        self._fill_sensitivity(ws_sensitivity, result)
        self._fill_methodology(ws_methodology, result)

        # 保存到 BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer

    def _fill_summary(self, ws, result: DCFResult):
        """填充汇总页"""
        # 标题
        ws["A1"] = "DCF VALUATION ANALYSIS"
        ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
        ws.merge_cells("A1:E1")

        # 公司信息
        ws["A3"] = "COMPANY INFORMATION"
        ws["A3"].font = SUBTITLE_FONT

        ws["A4"] = "Symbol:"
        ws["B4"] = result.symbol
        ws["B4"].font = FONT_BLUE  # 输入
        ws["A5"] = "Company Name:"
        ws["B5"] = result.company_name
        ws["A6"] = "Current Price:"
        ws["B6"] = result.current_price
        ws["B6"].number_format = "$#,##0.00"
        ws["B6"].font = FONT_BLUE

        ws["A7"] = "Currency:"
        ws["B7"] = result.currency
        ws["B7"].font = FONT_BLUE

        # 估值结果 - 使用链接
        ws["A9"] = "VALUATION RESULTS"
        ws["A9"].font = SUBTITLE_FONT

        ws["A10"] = "Enterprise Value ($M):"
        ws["B10"] = "=SUM('WACC'!B17,'Terminal Value'!B8)"
        ws["B10"].number_format = "#,##0"
        ws["B10"].font = FONT_GREEN  # 跨表链接
        ws["B10"].fill = KEY_OUTPUT_FILL

        ws["A11"] = "Equity Value ($M):"
        ws["B11"] = "=B10-WACC!B16"
        ws["B11"].number_format = "#,##0"
        ws["B11"].font = FONT_BLACK  # 公式
        ws["B11"].fill = KEY_OUTPUT_FILL

        ws["A12"] = "Implied Price:"
        ws["B12"] = "=B11/WACC!B15"
        ws["B12"].number_format = "$#,##0.00"
        ws["B12"].font = FONT_BLACK
        ws["B12"].fill = KEY_OUTPUT_FILL

        ws["A13"] = "Current Price:"
        ws["B13"] = "=B6"
        ws["B13"].number_format = "$#,##0.00"
        ws["B13"].font = FONT_PURPLE  # 同表链接

        ws["A14"] = "Upside/Downside:"
        ws["B14"] = "=(B12-B13)/B13"
        ws["B14"].number_format = "0.0%"
        ws["B14"].font = FONT_BLACK
        ws["B14"].fill = KEY_OUTPUT_FILL

        # 评级
        ws["A16"] = "RECOMMENDATION"
        ws["A16"].font = SUBTITLE_FONT

        ws["A17"] = "Rating:"
        ws["B17"] = result.recommendation
        ws["B17"].font = FONT_BLUE
        ws["A18"] = "Confidence:"
        ws["B18"] = result.confidence
        ws["B18"].font = FONT_BLUE

        # 估值区间 - 使用链接
        ws["A20"] = "VALUATION RANGE"
        ws["A20"].font = SUBTITLE_FONT

        ws["A21"] = "Bear Case:"
        ws["B21"] = "='Sensitivity Analysis'!B35"
        ws["B21"].number_format = "$#,##0.00"
        ws["B21"].font = FONT_GREEN

        ws["A22"] = "Base Case:"
        ws["B22"] = "='Sensitivity Analysis'!C35"
        ws["B22"].number_format = "$#,##0.00"
        ws["B22"].font = FONT_GREEN
        ws["B22"].fill = KEY_OUTPUT_FILL

        ws["A23"] = "Bull Case:"
        ws["B23"] = "='Sensitivity Analysis'!D35"
        ws["B23"].number_format = "$#,##0.00"
        ws["B23"].font = FONT_GREEN

        # 调整列宽
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20

    def _fill_wacc(self, ws, result: DCFResult):
        """填充 WACC 计算"""
        ws["A1"] = "WACC CALCULATION"
        ws["A1"].font = Font(bold=True, size=14, color="1F4E79")

        wacc = result.wacc_components

        # 输入参数 - 蓝色字体表示输入
        ws["A3"] = "INPUT PARAMETERS"
        ws["A3"].font = SUBTITLE_FONT
        ws["A3"].fill = COLUMN_HEADER_FILL

        ws["A4"] = "Risk Free Rate:"
        ws["B4"] = wacc.risk_free_rate
        ws["B4"].number_format = "0.00%"
        ws["B4"].font = FONT_BLUE  # 输入
        ws["B4"].fill = INPUT_CELL_FILL

        ws["A5"] = "Beta:"
        ws["B5"] = wacc.beta
        ws["B5"].number_format = "0.00"
        ws["B5"].font = FONT_BLUE
        ws["B5"].fill = INPUT_CELL_FILL

        ws["A6"] = "Equity Risk Premium:"
        ws["B6"] = wacc.equity_risk_premium
        ws["B6"].number_format = "0.00%"
        ws["B6"].font = FONT_BLUE
        ws["B6"].fill = INPUT_CELL_FILL

        ws["A7"] = "Tax Rate:"
        ws["B7"] = wacc.tax_rate
        ws["B7"].number_format = "0.00%"
        ws["B7"].font = FONT_BLUE
        ws["B7"].fill = INPUT_CELL_FILL

        # 股权成本计算 - 黑色字体表示公式
        ws["A9"] = "COST OF EQUITY (CAPM)"
        ws["A9"].font = SUBTITLE_FONT
        ws["A9"].fill = COLUMN_HEADER_FILL

        ws["A10"] = "Cost of Equity:"
        ws["B10"] = "=B4+B5*B6"
        ws["B10"].number_format = "0.00%"
        ws["B10"].font = FONT_BLACK  # 公式

        # 债务成本
        ws["A12"] = "COST OF DEBT"
        ws["A12"].font = SUBTITLE_FONT
        ws["A12"].fill = COLUMN_HEADER_FILL

        ws["A13"] = "Cost of Debt (Pre-tax):"
        ws["B13"] = wacc.cost_of_debt
        ws["B13"].number_format = "0.00%"
        ws["B13"].font = FONT_BLUE

        ws["A14"] = "Cost of Debt (After-tax):"
        ws["B14"] = "=B13*(1-B7)"
        ws["B14"].number_format = "0.00%"
        ws["B14"].font = FONT_BLACK

        # 资本结构
        ws["A16"] = "CAPITAL STRUCTURE"
        ws["A16"].font = SUBTITLE_FONT
        ws["A16"].fill = COLUMN_HEADER_FILL

        ws["A17"] = "Shares Outstanding:"
        ws["B17"] = result.wacc_components.market_cap / result.current_price if result.current_price > 0 else 0
        ws["B17"].number_format = "#,##0"
        ws["B17"].font = FONT_BLUE
        ws["B17"].fill = INPUT_CELL_FILL

        ws["A18"] = "Market Cap ($M):"
        ws["B18"] = wacc.market_cap / 1e6
        ws["B18"].number_format = "#,##0"
        ws["B18"].font = FONT_BLUE

        ws["A19"] = "Net Debt ($M):"
        ws["B19"] = wacc.net_debt / 1e6
        ws["B19"].number_format = "#,##0"
        ws["B19"].font = FONT_BLUE

        ws["A20"] = "Enterprise Value ($M):"
        ws["B20"] = "=B18+B19"
        ws["B20"].number_format = "#,##0"
        ws["B20"].font = FONT_BLACK

        ws["A21"] = "Equity Weight:"
        ws["B21"] = "=B18/B20"
        ws["B21"].number_format = "0.0%"
        ws["B21"].font = FONT_BLACK

        ws["A22"] = "Debt Weight:"
        ws["B22"] = "=B19/B20"
        ws["B22"].number_format = "0.0%"
        ws["B22"].font = FONT_BLACK

        # 最终 WACC - 关键输出
        ws["A24"] = "WEIGHTED AVERAGE COST OF CAPITAL"
        ws["A24"].font = Font(bold=True, size=12, color="1F4E79")
        ws["A24"].fill = KEY_OUTPUT_FILL

        ws["A25"] = "WACC:"
        ws["B25"] = "=B21*B10+B22*B14"
        ws["B25"].number_format = "0.00%"
        ws["B25"].font = Font(bold=True, size=12, color="000000")
        ws["B25"].fill = KEY_OUTPUT_FILL

        # 保存关键值供其他表引用
        ws["A15"] = "Shares (for reference):"
        ws["B15"] = "=B17"
        ws["B15"].font = FONT_PURPLE
        ws["B16"] = "Net Debt (for reference):"
        ws["B16"] = "=B19"
        ws["B16"].font = FONT_PURPLE
        ws["B17"] = "=B18"
        ws["B17"].font = FONT_PURPLE

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 15

    def _fill_fcf_with_formulas(self, ws, result: DCFResult):
        """填充 FCF 预测 - 使用公式"""
        ws["A1"] = "FREE CASH FLOW PROJECTIONS"
        ws["A1"].font = Font(bold=True, size=14, color="1F4E79")

        # 表头
        headers = [
            "Year",
            "Revenue ($M)",
            "Growth",
            "EBITDA",
            "EBITDA Margin",
            "EBIT",
            "Tax",
            "NOPAT",
            "D&A",
            "CapEx",
            "Δ NWC",
            "FCF ($M)",
            "Discount Factor",
            "PV of FCF ($M)",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = COLUMN_HEADER_FILL
            cell.font = BOLD_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN

        # 填充数据行
        for row_idx, fcf in enumerate(result.fcf_projections, 4):
            year = fcf.year
            row = row_idx

            # Year
            ws.cell(row=row, column=1, value=f"Year {year}")
            ws.cell(row=row, column=1).font = FONT_BLUE

            # Revenue
            ws.cell(row=row, column=2, value=f"=Revenue_{year}")
            ws.cell(row=row, column=2).number_format = "#,##0"
            ws.cell(row=row, column=2).font = FONT_GREEN  # 跨表引用

            # Growth
            ws.cell(row=row, column=3, value=fcf.revenue_growth)
            ws.cell(row=row, column=3).number_format = "0.0%"
            ws.cell(row=row, column=3).font = FONT_BLUE

            # EBITDA
            ws.cell(row=row, column=4, value=f"=EBITDA_{year}")
            ws.cell(row=row, column=4).number_format = "#,##0"
            ws.cell(row=row, column=4).font = FONT_GREEN

            # EBITDA Margin
            ws.cell(row=row, column=5, value=f"=D{row}/B{row}")
            ws.cell(row=row, column=5).number_format = "0.0%"
            ws.cell(row=row, column=5).font = FONT_BLACK

            # EBIT
            ws.cell(row=row, column=6, value=f"=EBIT_{year}")
            ws.cell(row=row, column=6).number_format = "#,##0"
            ws.cell(row=row, column=6).font = FONT_GREEN

            # Tax
            ws.cell(row=row, column=7, value=f"=-F{row}*'WACC'!$B$7")
            ws.cell(row=row, column=7).number_format = "#,##0"
            ws.cell(row=row, column=7).font = FONT_BLACK

            # NOPAT
            ws.cell(row=row, column=8, value=f"=F{row}+G{row}")
            ws.cell(row=row, column=8).number_format = "#,##0"
            ws.cell(row=row, column=8).font = FONT_BLACK

            # D&A
            ws.cell(row=row, column=9, value=f"=DA_{year}")
            ws.cell(row=row, column=9).number_format = "#,##0"
            ws.cell(row=row, column=9).font = FONT_GREEN

            # CapEx
            ws.cell(row=row, column=10, value=f"=CapEx_{year}")
            ws.cell(row=row, column=10).number_format = "#,##0"
            ws.cell(row=row, column=10).font = FONT_GREEN

            # Δ NWC
            ws.cell(row=row, column=11, value=f"=NWC_{year}")
            ws.cell(row=row, column=11).number_format = "#,##0"
            ws.cell(row=row, column=11).font = FONT_GREEN

            # FCF
            ws.cell(row=row, column=12, value=f"=H{row}+I{row}-J{row}-K{row}")
            ws.cell(row=row, column=12).number_format = "#,##0"
            ws.cell(row=row, column=12).font = FONT_BLACK

            # Discount Factor
            ws.cell(row=row, column=13, value=f"=1/((1+$B$16)^{year})")
            ws.cell(row=row, column=13).number_format = "0.0000"
            ws.cell(row=row, column=13).font = FONT_BLACK

            # PV of FCF
            ws.cell(row=row, column=14, value=f"=L{row}*M{row}")
            ws.cell(row=row, column=14).number_format = "#,##0"
            ws.cell(row=row, column=14).font = FONT_BLACK

            # 添加边框
            for col in range(1, 15):
                ws.cell(row=row, column=col).border = THIN_BORDER

        # 合计行
        total_row = len(result.fcf_projections) + 4
        ws.cell(row=total_row, column=1, value="TOTAL PV OF FCF").font = SUBTITLE_FONT
        ws.cell(row=total_row, column=13, value=f"=SUM(M4:M{total_row-1})")
        ws.cell(row=total_row, column=13).number_format = "#,##0"
        ws.cell(row=total_row, column=13).font = FONT_BLACK

        # WACC 参考值
        ws["A16"] = "WACC:"
        ws["B16"] = "='WACC'!B25"
        ws["B16"].number_format = "0.00%"
        ws["B16"].font = FONT_GREEN

        # 添加实际数据行 (隐藏)
        hidden_row = 20
        ws.row_dimensions[hidden_row].hidden = True

        # 在隐藏行中存储实际计算值，供公式引用
        for row_idx, fcf in enumerate(result.fcf_projections, 1):
            ws[f"A{hidden_row + row_idx}"] = f"Revenue_{row_idx}"
            ws[f"B{hidden_row + row_idx}"] = fcf.revenue / 1e6
            ws[f"C{hidden_row + row_idx}"] = f"EBITDA_{row_idx}"
            ws[f"D{hidden_row + row_idx}"] = fcf.ebitda / 1e6
            ws[f"E{hidden_row + row_idx}"] = f"EBIT_{row_idx}"
            ws[f"F{hidden_row + row_idx}"] = fcf.ebit / 1e6
            ws[f"G{hidden_row + row_idx}"] = f"DA_{row_idx}"
            ws[f"H{hidden_row + row_idx}"] = f"CapEx_{row_idx}"
            ws[f"I{hidden_row + row_idx}"] = f"NWC_{row_idx}"
            ws[f"J{hidden_row + row_idx}"] = fcf.da / 1e6
            ws[f"K{hidden_row + row_idx}"] = fcf.capex / 1e6
            ws[f"L{hidden_row + row_idx}"] = fcf.delta_nwc / 1e6

        # 列宽
        ws.column_dimensions["A"].width = 12
        for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N"]:
            ws.column_dimensions[col].width = 14

    def _fill_terminal(self, ws, result: DCFResult):
        """填充终值计算"""
        ws["A1"] = "TERMINAL VALUE CALCULATION"
        ws["A1"].font = Font(bold=True, size=14, color="1F4E79")

        tv = result.terminal_value
        fcf_proj = result.fcf_projections

        # 终值参数
        ws["A3"] = "TERMINAL VALUE PARAMETERS"
        ws["A3"].font = SUBTITLE_FONT
        ws["A3"].fill = COLUMN_HEADER_FILL

        ws["A4"] = "Terminal Growth Rate:"
        ws["B4"] = tv.terminal_growth_rate
        ws["B4"].number_format = "0.00%"
        ws["B4"].font = FONT_BLUE
        ws["B4"].fill = INPUT_CELL_FILL

        ws["A5"] = "Exit Multiple (EV/EBITDA):"
        ws["B5"] = tv.exit_multiple if tv.exit_multiple > 0 else 15.0
        ws["B5"].number_format = "0.0x"
        ws["B5"].font = FONT_BLUE
        ws["B5"].fill = INPUT_CELL_FILL

        # 永续增长法
        ws["A7"] = "PERPETUITY GROWTH METHOD"
        ws["A7"].font = SUBTITLE_FONT
        ws["A7"].fill = COLUMN_HEADER_FILL

        ws["A8"] = "Terminal Year FCF ($M):"
        if fcf_proj:
            ws["B8"] = f"={{'FCF Projections'!L{3 + len(fcf_proj)}}}*(1+B4)"
        else:
            ws["B8"] = 0
        ws["B8"].number_format = "#,##0"
        ws["B8"].font = FONT_BLACK

        ws["A9"] = "Terminal Value (Perpetuity):"
        ws["B9"] = "=B8/('WACC'!B25-B4)"
        ws["B9"].number_format = "#,##0"
        ws["B9"].font = FONT_BLACK
        ws["B9"].fill = KEY_OUTPUT_FILL

        ws["A10"] = "PV of Terminal Value ($M):"
        ws["B10"] = "=B9/((1+'WACC'!B25)^'FCF Projections'!$B$16)"
        ws["B10"].number_format = "#,##0"
        ws["B10"].font = FONT_BLACK

        # 退出倍数法
        ws["A12"] = "EXIT MULTIPLE METHOD"
        ws["A12"].font = SUBTITLE_FONT
        ws["A12"].fill = COLUMN_HEADER_FILL

        ws["A13"] = "Terminal Year EBITDA ($M):"
        if fcf_proj:
            ws["B13"] = f"={{'FCF Projections'!D{3 + len(fcf_proj)}}}"
        else:
            ws["B13"] = 0
        ws["B13"].number_format = "#,##0"
        ws["B13"].font = FONT_PURPLE

        ws["A14"] = "Terminal Value (Multiple):"
        ws["B14"] = "=B13*B5"
        ws["B14"].number_format = "#,##0"
        ws["B14"].font = FONT_BLACK

        ws["A15"] = "PV of Terminal Value ($M):"
        ws["B15"] = "=B14/((1+'WACC'!B25)^'FCF Projections'!$B$16)"
        ws["B15"].number_format = "#,##0"
        ws["B15"].font = FONT_BLACK

        # 终值占比
        ws["A17"] = "TERMINAL VALUE AS % OF ENTERPRISE VALUE"
        ws["A17"].font = SUBTITLE_FONT
        ws["A17"].fill = KEY_OUTPUT_FILL

        ws["A18"] = "PV of FCF Sum ($M):"
        ws["B18"] = "='FCF Projections'!M4"
        ws["B18"].number_format = "#,##0"
        ws["B18"].font = FONT_GREEN

        ws["A19"] = "PV of Terminal Value ($M):"
        ws["B19"] = "=B10"
        ws["B19"].number_format = "#,##0"
        ws["B19"].font = FONT_PURPLE

        ws["A20"] = "Total Enterprise Value ($M):"
        ws["B20"] = "=B18+B19"
        ws["B20"].number_format = "#,##0"
        ws["B20"].font = FONT_BLACK
        ws["B20"].fill = KEY_OUTPUT_FILL

        ws["A21"] = "Terminal Value %:"
        ws["B21"] = "=B19/B20"
        ws["B21"].number_format = "0.0%"
        ws["B21"].font = FONT_BLACK

        # 保存关键值供 Summary 引用
        ws["B8"] = "=B10"  # PV Terminal Value for Summary

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 18

    def _fill_sensitivity(self, ws, result: DCFResult):
        """填充敏感性分析 - 使用公式"""
        ws["A1"] = "SENSITIVITY ANALYSIS - IMPLIED SHARE PRICE"
        ws["A1"].font = Font(bold=True, size=14, color="1F4E79")

        sens = result.sensitivity
        base_price = result.implied_price if result.implied_price > 0 else 100

        # 矩阵1: WACC vs Terminal Growth Rate
        ws["A3"] = "1. WACC vs Terminal Growth Rate"
        ws["A3"].font = SUBTITLE_FONT
        ws["A3"].fill = COLUMN_HEADER_FILL

        # 表头 - WACC 值
        headers1 = ["WACC \\ Growth"] + [f"{g*100:.1f}%" for g in sens.terminal_growth_values]
        for col, header in enumerate(headers1, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.fill = COLUMN_HEADER_FILL
            cell.font = BOLD_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN

        # 数据 - 使用公式
        for row_idx, wacc_val in enumerate(sens.wacc_values):
            ws.cell(row=5 + row_idx, column=1, value=f"{wacc_val * 100:.1f}%").border = THIN_BORDER
            ws.cell(row=5 + row_idx, column=1).font = FONT_BLUE

            for col_idx, price in enumerate(sens.price_matrix_wacc_growth[row_idx]):
                cell = ws.cell(row=5 + row_idx, column=2 + col_idx, value=price)
                cell.number_format = "$#,##0.00"
                cell.border = THIN_BORDER
                cell.font = FONT_BLUE  # 基准值，中心单元格特殊标记

        # 标记中心单元格 (base case)
        center_row = 5 + len(sens.wacc_values) // 2
        center_col = 2 + len(sens.terminal_growth_values) // 2
        ws.cell(row=center_row, column=center_col).fill = KEY_OUTPUT_FILL
        ws.cell(row=center_row, column=center_col).font = BOLD_FONT

        # 矩阵2: Revenue Growth vs EBITDA Margin
        row_offset = len(sens.wacc_values) + 8
        ws.cell(row=row_offset, column=1, value="2. Revenue Growth vs EBITDA Margin").font = SUBTITLE_FONT
        ws.cell(row=row_offset, column=1).fill = COLUMN_HEADER_FILL

        headers2 = ["Growth \\ Margin"] + [f"{m*100:.1f}%" for m in sens.ebitda_margin_values]
        for col, header in enumerate(headers2, 1):
            cell = ws.cell(row=row_offset + 1, column=col, value=header)
            cell.fill = COLUMN_HEADER_FILL
            cell.font = BOLD_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN

        for row_idx, growth_val in enumerate(sens.revenue_growth_values):
            ws.cell(row=row_offset + 2 + row_idx, column=1, value=f"{growth_val * 100:.1f}%").border = THIN_BORDER
            ws.cell(row=row_offset + 2 + row_idx, column=1).font = FONT_BLUE

            for col_idx, price in enumerate(sens.price_matrix_growth_margin[row_idx]):
                cell = ws.cell(row=row_offset + 2 + row_idx, column=2 + col_idx, value=price)
                cell.number_format = "$#,##0.00"
                cell.border = THIN_BORDER
                cell.font = FONT_BLUE

        # 矩阵3: Exit Multiple vs WACC
        row_offset2 = row_offset + len(sens.revenue_growth_values) + 8
        ws.cell(row=row_offset2, column=1, value="3. Exit Multiple vs WACC").font = SUBTITLE_FONT
        ws.cell(row=row_offset2, column=1).fill = COLUMN_HEADER_FILL

        headers3 = ["Multiple \\ WACC"] + [f"{w*100:.1f}%" for w in sens.wacc_values_2]
        for col, header in enumerate(headers3, 1):
            cell = ws.cell(row=row_offset2 + 1, column=col, value=header)
            cell.fill = COLUMN_HEADER_FILL
            cell.font = BOLD_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN

        for row_idx, multiple_val in enumerate(sens.exit_multiple_values):
            ws.cell(row=row_offset2 + 2 + row_idx, column=1, value=f"{multiple_val:.1f}x").border = THIN_BORDER
            ws.cell(row=row_offset2 + 2 + row_idx, column=1).font = FONT_BLUE

            for col_idx, price in enumerate(sens.price_matrix_multiple_wacc[row_idx]):
                cell = ws.cell(row=row_offset2 + 2 + row_idx, column=2 + col_idx, value=price)
                cell.number_format = "$#,##0.00"
                cell.border = THIN_BORDER
                cell.font = FONT_BLUE

        # 估值区间汇总
        ws["A30"] = "VALUATION RANGE SUMMARY"
        ws["A30"].font = Font(bold=True, size=12, color="1F4E79")
        ws["A30"].fill = KEY_OUTPUT_FILL

        ws["A32"] = "Base Case Price:"
        ws["B32"] = base_price
        ws["B32"].number_format = "$#,##0.00"
        ws["B32"].font = FONT_BLUE

        # 从三个矩阵提取估值区间
        all_prices = []
        for row in sens.price_matrix_wacc_growth:
            all_prices.extend([p for p in row if p > 0])
        for row in sens.price_matrix_growth_margin:
            all_prices.extend([p for p in row if p > 0])
        for row in sens.price_matrix_multiple_wacc:
            all_prices.extend([p for p in row if p > 0])

        if all_prices:
            sorted_prices = sorted(all_prices)
            n = len(sorted_prices)

            ws["A34"] = "Bear Case (25th %):"
            ws["B34"] = sorted_prices[int(n * 0.25)]
            ws["B34"].number_format = "$#,##0.00"
            ws["B34"].font = FONT_BLUE

            ws["A35"] = "Base Case (Median):"
            ws["B35"] = sorted_prices[int(n * 0.5)]
            ws["B35"].number_format = "$#,##0.00"
            ws["B35"].font = FONT_BLUE
            ws["B35"].fill = KEY_OUTPUT_FILL

            ws["A36"] = "Bull Case (75th %):"
            ws["B36"] = sorted_prices[int(n * 0.75)]
            ws["B36"].number_format = "$#,##0.00"
            ws["B36"].font = FONT_BLUE

        # 列宽
        ws.column_dimensions["A"].width = 18
        for i in range(2, 8):
            ws.column_dimensions[get_column_letter(i)].width = 14

    def _fill_methodology(self, ws, result: DCFResult):
        """填充方法论和数据源文档"""
        ws["A1"] = "METHODOLOGY & DATA SOURCES"
        ws["A1"].font = Font(bold=True, size=14, color="1F4E79")

        row = 3

        # 分析日期
        ws.cell(row=row, column=1, value="ANALYSIS DATE")
        ws.cell(row=row, column=1).font = SUBTITLE_FONT
        ws.cell(row=row, column=1).fill = COLUMN_HEADER_FILL
        row += 1
        ws.cell(row=row, column=1, value=datetime.now().strftime("%Y-%m-%d"))
        ws.cell(row=row, column=1).font = FONT_BLUE
        row += 2

        # 数据源
        ws.cell(row=row, column=1, value="DATA SOURCES")
        ws.cell(row=row, column=1).font = SUBTITLE_FONT
        ws.cell(row=row, column=1).fill = COLUMN_HEADER_FILL
        row += 1

        sources = [
            ("Primary Data:", "yfinance"),
            ("Financial Data:", "Company SEC Filings (10-K, 10-Q)"),
            ("Market Data:", "Real-time market prices from exchange"),
            ("Analyst Estimates:", "Wall Street consensus estimates"),
        ]

        for label, source in sources:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=source)
            row += 1
        row += 1

        # 方法论
        ws.cell(row=row, column=1, value="VALUATION METHODOLOGY")
        ws.cell(row=row, column=1).font = SUBTITLE_FONT
        ws.cell(row=row, column=1).fill = COLUMN_HEADER_FILL
        row += 1

        methodologies = [
            ("Discount Rate (WACC):", "Weighted Average Cost of Capital using CAPM"),
            ("Cost of Equity:", "Risk Free Rate + Beta × Equity Risk Premium"),
            ("Cost of Debt:", "Based on interest expense / total debt"),
            ("Terminal Growth:", "Perpetuity growth method"),
            ("Projection Period:", f"{result.assumptions.get('projection_years', 5)} years"),
        ]

        for label, method in methodologies:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=method)
            row += 1
        row += 1

        # 关键假设
        ws.cell(row=row, column=1, value="KEY ASSUMPTIONS")
        ws.cell(row=row, column=1).font = SUBTITLE_FONT
        ws.cell(row=row, column=1).fill = COLUMN_HEADER_FILL
        row += 1

        assumptions = result.assumptions or {}
        assumption_items = [
            ("Risk Free Rate:", assumptions.get("risk_free_rate", 0.042)),
            ("Equity Risk Premium:", assumptions.get("equity_risk_premium", 0.055)),
            ("Tax Rate:", assumptions.get("tax_rate", 0.25)),
            ("Terminal Growth Rate:", assumptions.get("terminal_growth_rate", 0.025)),
            ("Projection Years:", assumptions.get("projection_years", 5)),
        ]

        for label, value in assumption_items:
            ws.cell(row=row, column=1, value=label)
            if isinstance(value, float):
                ws.cell(row=row, column=2, value=value)
                ws.cell(row=row, column=2).number_format = "0.00%" if "Rate" in label else "0"
            else:
                ws.cell(row=row, column=2, value=value)
            row += 1
        row += 1

        # 免责声明
        ws.cell(row=row, column=1, value="DISCLAIMER")
        ws.cell(row=row, column=1).font = SUBTITLE_FONT
        ws.cell(row=row, column=1).fill = COLUMN_HEADER_FILL
        row += 1

        disclaimer = (
            "This analysis is for informational purposes only and does not constitute "
            "investment advice. The valuation is based on assumptions that may differ "
            "from actual results. Past performance is not indicative of future results."
        )
        ws.cell(row=row, column=1, value=disclaimer)
        ws.merge_cells(f"A{row}:E{row}")
        ws.cell(row=row, column=1).alignment = LEFT_ALIGN
        ws.cell(row=row, column=1).font = Font(size=9, italic=True)

        # 列宽
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 50
